from django.http import HttpResponse
from django.utils import timezone as djtz
from django.utils.translation import gettext  # non-lazy
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from orders.models import Order


def _ensure_local_aware(dt):
    if dt is None:
        return None
    if djtz.is_naive(dt):
        return djtz.make_aware(dt, djtz.get_current_timezone())
    return dt


def export_orders_to_excel(start_dt, end_dt):
    """
    Export Orders between start_dt and end_dt to an XLSX HttpResponse.
    Ensure all values passed to openpyxl are real strings (not lazy proxies).
    """
    qs = Order.objects.filter(created_at__gte=start_dt, created_at__lte=end_dt).order_by('created_at')

    wb = Workbook()
    ws = wb.active
    # Use gettext (non-lazy) to produce real strings
    ws.title = gettext("Orders")

    # Columns: use gettext for header labels so they're real strings
    columns = [
        (gettext("ID"), "id"),
        (gettext("Created at"), "created_at"),
        (gettext("Employee (id)"), "employee_id"),
        (gettext("Employee Name"), "employee_name"),
        (gettext("Food size"), "food_size_display"),
        (gettext("Is cancelled"), "is_cancelled"),
    ]

    # Header row (force str() just in case)
    for idx, (col_name, _) in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=str(col_name))

    # Rows
    for row_idx, order in enumerate(qs, start=2):
        values = {
            "id": order.id,
            "created_at": djtz.localtime(order.created_at).strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "",
            "employee_id": order.employee.id if order.employee else "",
            "employee_name": order.name if order.employee else "",
            "food_size_display": order.get_food_size_display(),
            "is_cancelled": gettext("Yes") if bool(order.is_cancelled) else gettext("No"),
        }

        for col_idx, (_, key) in enumerate(columns, start=1):
            cell_value = values.get(key, "")
            # ensure cell_value is a plain python value (str, int, bool...)
            # convert lazy proxies to str (defensive)
            if hasattr(cell_value, "__class__") and getattr(cell_value, "__class__").__name__ == "Promise":
                cell_value = str(cell_value)
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Auto-size columns
    for i in range(1, len(columns) + 1):
        col_letter = get_column_letter(i)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    # Build response: use gettext for the 'orders' portion, dates left as ISO strings
    start_str = start_dt.strftime("%Y%m%d")
    end_str = end_dt.strftime("%Y%m%d")
    filename = f"{gettext('orders')}_{start_str}_{end_str}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
