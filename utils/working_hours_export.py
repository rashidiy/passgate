# utils/working_hours_export.py
from __future__ import annotations

from calendar import monthrange, month_name as EN_MONTH_NAME
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, time, date
from typing import Dict, Iterable, List, Optional, Tuple

from django.http import HttpResponse
from django.utils import timezone as djtz
from django.utils.encoding import force_str
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from devices.models import Device, Event
from employees.models import Employee, WorkingHourException

# ──────────────────────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────────────────────
DAY_START = time(8, 0, 0)   # 08:00
NIGHT_START = time(20, 0, 0)  # 20:00

EXC_LABEL_LABOR  = _("Labor Leave")
EXC_LABEL_UNPAID = _("Unpaid Leave")
EXC_LABEL_TRIP   = _("Working Trip")

# Two-letter control codes (DO NOT translate; used in Excel number formats)
EXC_CODE_LABOR  = "LL"  # LaborLib
EXC_CODE_UNPAID = "UL"  # UnpaidLib
EXC_CODE_TRIP   = "WT"  # WorkingTribute (not shown in visible cells)

# Weekend fill (yellow) – use 8-digit ARGB for consistency
WEEKEND_FILL = PatternFill(fill_type="solid", start_color="00FFFF00", end_color="00FFFF00")

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
def s(v) -> str:
    return force_str(v)

def _tz():
    return djtz.get_default_timezone()

def _aware(dt: datetime) -> datetime:
    return djtz.make_aware(dt, _tz()) if djtz.is_naive(dt) else dt.astimezone(_tz())

@dataclass(frozen=True)
class Interval:
    start: datetime
    end: datetime

def _overlap_seconds(a: Interval, b: Interval) -> float:
    start = max(a.start, b.start)
    end = min(a.end, b.end)
    if end <= start:
        return 0.0
    return (end - start).total_seconds()

def _daterange_days(start: datetime, end: datetime) -> Iterable[date]:
    cur = start.astimezone(_tz()).date()
    last = end.astimezone(_tz()).date()
    while cur <= last:
        yield cur
        cur += timedelta(days=1)

def _month_edges(year: int, month: int) -> Interval:
    last = monthrange(year, month)[1]
    return Interval(
        _aware(datetime(year, month, 1, 0, 0, 0)),
        _aware(datetime(year, month, last, 23, 59, 59)),
    )

def _day_night_windows(d: date) -> Tuple[Interval, Interval, Interval]:
    d0 = _aware(datetime.combine(d, DAY_START))
    d1 = _aware(datetime.combine(d, NIGHT_START))
    d_end = _aware(datetime.combine(d, time(23, 59, 59)))
    next_day = d + timedelta(days=1)
    n2_from = _aware(datetime.combine(next_day, time.min))
    n2_to = _aware(datetime.combine(next_day, DAY_START))
    return Interval(d0, d1), Interval(d1, d_end), Interval(n2_from, n2_to)

def _fmt_hhmm(seconds: int) -> str:
    minutes = int(round(seconds / 60))
    h, m = divmod(minutes, 60)
    return f"{h}:{m:02d}"

# ──────────────────────────────────────────────────────────────────────────────
# Interval utilities
# ──────────────────────────────────────────────────────────────────────────────
def _pair_presence(events: Iterable[Event]) -> List[Interval]:
    """ENTER→EXIT pairs from devices (onsite presence)."""
    pairs: List[Interval] = []
    opened: List[datetime] = []
    for ev in events:
        if not ev.device:
            continue
        if ev.device.type == Device.DeviceTypes.ENTER:
            opened.append(ev.timestamp)
        elif ev.device.type == Device.DeviceTypes.EXIT:
            if opened:
                start = opened.pop(0)
                end = ev.timestamp
                if end > start:
                    pairs.append(Interval(_aware(start), _aware(end)))
    return pairs

def _trip_intervals(employee: Optional[Employee], q_start: datetime, q_end: datetime) -> List[Interval]:
    """WorkingTrip (WT) intervals from exceptions, clamped to [q_start, q_end]."""
    if not employee:
        return []
    qs = WorkingHourException.objects.filter(
        employee=employee,
        start_time__lte=q_end,
        end_time__gte=q_start,
        exception_type=WorkingHourException.ExceptionTypes.WorkingTrip,
    ).values_list("start_time", "end_time")
    out: List[Interval] = []
    for sdt, edt in qs:
        sdt = _aware(sdt); edt = _aware(edt)
        sdt = max(sdt, q_start); edt = min(edt, q_end)
        if edt > sdt:
            out.append(Interval(sdt, edt))
    return out

def _merge_intervals(intervals: List[Interval]) -> List[Interval]:
    """Union overlapping/touching intervals to avoid double-counting."""
    if not intervals:
        return []
    arr = sorted(intervals, key=lambda iv: iv.start)
    merged: List[Interval] = []
    cur = arr[0]
    for iv in arr[1:]:
        if iv.start <= cur.end:
            cur = Interval(cur.start, max(cur.end, iv.end))
        else:
            merged.append(cur)
            cur = iv
    merged.append(cur)
    return merged

# ──────────────────────────────────────────────────────────────────────────────
# Exceptions flags
# ──────────────────────────────────────────────────────────────────────────────
@dataclass
class DayExceptions:
    has_labor: bool = False
    has_unpaid: bool = False
    has_trip: bool = False

    def labels(self) -> List[str]:
        out = []
        if self.has_labor:  out.append(s(EXC_LABEL_LABOR))
        if self.has_unpaid: out.append(s(EXC_LABEL_UNPAID))
        if self.has_trip:   out.append(s(EXC_LABEL_TRIP))
        return out

    def display_codes(self) -> List[str]:
        """Codes to DISPLAY in cells (WT excluded; WT contributes numeric time)."""
        codes: List[str] = []
        if self.has_labor:  codes.append(EXC_CODE_LABOR)
        if self.has_unpaid: codes.append(EXC_CODE_UNPAID)
        return codes

def _exceptions_by_day(employee: Employee, year: int, month: int) -> Dict[int, DayExceptions]:
    days = monthrange(year, month)[1]
    first = _aware(datetime(year, month, 1, 0, 0, 0))
    last  = _aware(datetime(year, month, days, 23, 59, 59))
    excs = WorkingHourException.objects.filter(
        employee=employee,
        start_time__lte=last,
        end_time__gte=first,
    )
    by_day: Dict[int, DayExceptions] = defaultdict(DayExceptions)
    for e in excs:
        sdt = _aware(e.start_time)
        edt = _aware(e.end_time)
        cur = sdt.date()
        while cur <= edt.date():
            if cur.year == year and cur.month == month:
                flags = by_day[cur.day]
                if e.exception_type == WorkingHourException.ExceptionTypes.LaborLeave:
                    flags.has_labor = True
                elif e.exception_type == WorkingHourException.ExceptionTypes.UnpaidLeave:
                    flags.has_unpaid = True
                elif e.exception_type == WorkingHourException.ExceptionTypes.WorkingTrip:
                    flags.has_trip = True
            cur += timedelta(days=1)
    return by_day

# ──────────────────────────────────────────────────────────────────────────────
# Split presence into per-day Day/Night SECONDS
# ──────────────────────────────────────────────────────────────────────────────
def _split_presence_seconds(pairs: List[Interval]) -> Dict[date, Tuple[int, int]]:
    result: Dict[date, Tuple[int, int]] = defaultdict(lambda: (0, 0))
    for iv in pairs:
        if iv.end <= iv.start:
            continue
        for d in _daterange_days(iv.start, iv.end):
            day_win, n1_win, n2_win = _day_night_windows(d)
            day_sec = int(round(_overlap_seconds(iv, day_win)))
            n1_sec  = int(round(_overlap_seconds(iv, n1_win)))
            n2_sec  = int(round(_overlap_seconds(iv, n2_win)))
            prev_d, prev_n = result[d]
            result[d] = (prev_d + day_sec, prev_n + n1_sec + n2_sec)
    return result

# ──────────────────────────────────────────────────────────────────────────────
# Export API
# ──────────────────────────────────────────────────────────────────────────────
@dataclass
class EmpRow:
    emp_id: Optional[int]   # None for orphan/deleted
    name: str
    real: Optional[Employee]  # None for orphan

def _iter_year_months_in_range(start_dt: datetime, end_dt: datetime) -> List[Tuple[int, int]]:
    """Inclusive list of (year, month) that overlap [start_dt, end_dt]."""
    cur = datetime(start_dt.year, start_dt.month, 1, tzinfo=_tz())
    end = datetime(end_dt.year, end_dt.month, 1, tzinfo=_tz())
    out: List[Tuple[int, int]] = []
    while cur <= end:
        out.append((cur.year, cur.month))
        cur = datetime(cur.year + 1, 1, 1, tzinfo=_tz()) if cur.month == 12 else datetime(cur.year, cur.month + 1, 1, tzinfo=_tz())
    return out

def export_working_hours_to_excel(
    year: Optional[int] = None,
    months: Optional[List[int]] = None,
    include_all_employees: bool = False,
    start_dt: Optional[datetime] = None,
    end_dt: Optional[datetime] = None,
) -> HttpResponse:
    """
    Build an XLSX with one sheet per month.

    Modes:
      - (year[, months])  → traditional year/month export
      - (start_dt, end_dt) → arbitrary interval, auto-sliced to months

    Rules:
      - Cells are NUMERIC hours (=seconds/3600).
      - **LL/UL**: show initials but numeric value is **0** → excluded from totals.
      - **WT**: contributes numeric time (no initials), comment shows “+ WT hh:mm”.
      - Weekends (Sat/Sun) highlighted yellow.
    """
    wb = Workbook()

    # Hidden mirror sheet for machine-readable flags (first sheet)
    flags_ws = wb.active
    flags_ws.title = "__flags_internal"
    flags_ws.sheet_state = "hidden"
    flags_ws["A1"].value = "Legend: LL=LaborLib, UL=UnpaidLib, WT=<seconds> (added work)"

    # Figure months
    ym_list: List[Tuple[int, int]]
    range_bounds: Optional[Tuple[datetime, datetime]] = None
    if start_dt and end_dt:
        start_dt = _aware(start_dt); end_dt = _aware(end_dt)
        if end_dt < start_dt:
            start_dt, end_dt = end_dt, start_dt
        ym_list = _iter_year_months_in_range(start_dt, end_dt)
        range_bounds = (start_dt, end_dt)
    else:
        assert year is not None, "Provide either (year[, months]) or (start_dt, end_dt)."
        if months is None:
            month_qs = (Event.objects.filter(timestamp__year=year)
                        .values_list("timestamp__month", flat=True)
                        .distinct())
            months = sorted(int(m) for m in month_qs)
        ym_list = [(year, m) for m in months]

    if not ym_list:
        ws = wb.create_sheet("Temp")
        ws.sheet_state = "visible"
        ws.title = f"{(year or _aware(djtz.now()).year)} {s(_(EN_MONTH_NAME[1]))}"

    for (y, m) in ym_list:
        days_in_month = monthrange(y, m)[1]
        month_iv = _month_edges(y, m)

        # Clamp to requested interval
        q_start, q_end = month_iv.start, month_iv.end
        if range_bounds:
            q_start = max(q_start, range_bounds[0])
            q_end   = min(q_end,   range_bounds[1])
            if q_end <= q_start:
                continue

        # Events in range
        ev_list = list(
            Event.objects
            .filter(timestamp__gte=q_start, timestamp__lte=q_end)
            .select_related("employee", "device")
            .order_by("employee_id", "timestamp", "serial_no")
        )

        # Build target rows
        rows: List[EmpRow] = []
        if include_all_employees:
            for emp in Employee.objects.all().order_by("name"):
                rows.append(EmpRow(emp_id=emp.id, name=emp.name, real=emp))
        else:
            emp_ids = (Event.objects
                       .filter(timestamp__gte=q_start, timestamp__lte=q_end)
                       .exclude(employee__isnull=True)
                       .values_list("employee_id", flat=True)
                       .distinct())
            for emp in Employee.objects.filter(id__in=emp_ids).order_by("name"):
                rows.append(EmpRow(emp_id=emp.id, name=emp.name, real=emp))
            orphan_names = (Event.objects
                            .filter(timestamp__gte=q_start, timestamp__lte=q_end, employee__isnull=True)
                            .values_list("employee_name", flat=True)
                            .distinct())
            for oname in orphan_names:
                if oname:
                    rows.append(EmpRow(emp_id=None, name=oname, real=None))

        rows.sort(key=lambda r: (r.emp_id is None, r.emp_id or 0, r.name.lower()))

        # Visible month sheet (ensure visible)
        title = f"{y} {s(_(EN_MONTH_NAME[m]))}"
        ws = wb.create_sheet(title)
        ws.sheet_state = "visible"

        # Per-month hidden flags sheet
        flags_month_ws = wb.create_sheet(f"flags_{y}_{m:02d}")
        flags_month_ws.sheet_state = "hidden"

        # Headers
        header_row_1 = [s(_("Employee ID")), s(_("Employee"))]
        header_row_2 = ["", ""]

        def day_col(d: int) -> int:
            return 3 + (d - 1) * 2

        for d in range(1, days_in_month + 1):
            header_row_1 += [str(d), ""]
            header_row_2 += [s(_("Day")), s(_("Night"))]

        header_row_1 += ["", "", "", s(_("Exceptions"))]
        header_row_2 += [s(_("Day Total")), s(_("Night Total")), s(_("Total Hours")), ""]

        ws.append([s(x) for x in header_row_1])
        ws.append([s(x) for x in header_row_2])

        # Merge header cells
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # Employee ID
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # Employee
        for d in range(1, days_in_month + 1):
            c1 = day_col(d)
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1 + 1)

        total_day_col   = day_col(days_in_month) + 2
        total_night_col = total_day_col + 1
        total_col       = total_night_col + 1
        exc_col         = total_col + 1

        # Widths
        ws.column_dimensions[get_column_letter(1)].width = 14  # Employee ID
        ws.column_dimensions[get_column_letter(2)].width = 26  # Employee
        for d in range(1, days_in_month + 1):
            ws.column_dimensions[get_column_letter(day_col(d))].width     = 8
            ws.column_dimensions[get_column_letter(day_col(d) + 1)].width = 8
        ws.column_dimensions[get_column_letter(total_day_col)].width   = 12
        ws.column_dimensions[get_column_letter(total_night_col)].width = 13
        ws.column_dimensions[get_column_letter(total_col)].width       = 12
        ws.column_dimensions[get_column_letter(exc_col)].width         = 24

        # Weekend highlight on headers (Sat=5, Sun=6)
        for d in range(1, days_in_month + 1):
            weekday = date(y, m, d).weekday()
            if weekday >= 5:
                c1 = day_col(d)
                for rr in (1, 2):
                    ws.cell(row=rr, column=c1).fill     = WEEKEND_FILL
                    ws.cell(row=rr, column=c1 + 1).fill = WEEKEND_FILL

        # Rows
        for rowinfo in rows:
            if rowinfo.real is not None:
                evs = [e for e in ev_list if e.employee_id == rowinfo.emp_id]
            else:
                evs = [e for e in ev_list if e.employee_id is None and e.employee_name == rowinfo.name]

            pairs         = _pair_presence(evs)
            wt_intervals  = _trip_intervals(rowinfo.real, q_start, q_end)
            all_intervals = _merge_intervals(pairs + wt_intervals)

            split_sec  = _split_presence_seconds(all_intervals)
            wt_split   = _split_presence_seconds(wt_intervals)
            exc_by_day = _exceptions_by_day(rowinfo.real, y, m) if rowinfo.real else {}

            base = [(rowinfo.emp_id if rowinfo.emp_id is not None else ""), rowinfo.name]
            for _i in range(days_in_month):
                base += [0, 0]
            base += ["", "", "", ""]
            ws.append(base)
            r = ws.max_row

            # Mirror row header to flags sheet
            flags_month_ws.append([base[0], base[1]] + [""] * (days_in_month * 2 + 4))

            for d in range(1, days_in_month + 1):
                c_day = day_col(d)
                c_nig = c_day + 1
                dt_key = date(y, m, d)

                day_sec = night_sec = 0
                if dt_key in split_sec:
                    day_sec, night_sec = split_sec[dt_key]

                # Default numeric values (hours)
                day_cell   = ws.cell(row=r, column=c_day);   day_cell.value   = f"={day_sec}/3600";   day_cell.number_format   = "0.00"
                night_cell = ws.cell(row=r, column=c_nig);   night_cell.value = f"={night_sec}/3600"; night_cell.number_format = "0.00"

                # Weekend fill
                if dt_key.weekday() >= 5:
                    day_cell.fill = WEEKEND_FILL
                    night_cell.fill = WEEKEND_FILL

                # Flags / display logic
                code_str = ""
                wt_day_added = wt_night_added = 0
                if rowinfo.real is not None:
                    flags = exc_by_day.get(d)
                    if flags:
                        # If LL or UL — treat as NON-working: force numeric 0, display initials
                        display_codes = flags.display_codes()  # LL/UL only
                        if flags.has_labor or flags.has_unpaid:
                            # force zeros, still show codes
                            day_cell.value = 0
                            night_cell.value = 0
                            if display_codes:
                                code_str = ",".join(display_codes)
                                suffix = f' "{code_str}"'
                                day_cell.number_format   = "0.00" + suffix
                                night_cell.number_format = "0.00" + suffix
                        else:
                            # No LL/UL: if there ARE display codes (unlikely) just add suffix
                            if display_codes:
                                code_str = ",".join(display_codes)
                                suffix = f' "{code_str}"'
                                day_cell.number_format   = "0.00" + suffix
                                night_cell.number_format = "0.00" + suffix

                        # WT – contributes numeric time, add a comment only
                        if flags.has_trip and dt_key in wt_split:
                            wt_day_added, wt_night_added = wt_split[dt_key]
                            parts = []
                            if wt_day_added > 0:
                                parts.append(f"Day + WT {_fmt_hhmm(wt_day_added)}")
                            if wt_night_added > 0:
                                parts.append(f"Night + WT {_fmt_hhmm(wt_night_added)}")
                            if parts:
                                note_text = "; ".join(parts)
                                author = s(_("System"))
                                if not day_cell.comment:
                                    day_cell.comment = Comment(note_text, author)
                                if not night_cell.comment:
                                    night_cell.comment = Comment(note_text, author)

                # Hidden flags sheet:
                #  - write LL/UL codes
                #  - write WT seconds as WT=<seconds> (per cell) for scripting later
                flags_month_ws.cell(row=r, column=c_day).value = code_str or ""
                flags_month_ws.cell(row=r, column=c_nig).value = code_str or ""
                if wt_day_added > 0:
                    flags_month_ws.cell(row=r, column=c_day).value += ("," if code_str else "") + f"{EXC_CODE_TRIP}={wt_day_added}"
                if wt_night_added > 0:
                    prefix = flags_month_ws.cell(row=r, column=c_nig).value or ""
                    flags_month_ws.cell(row=r, column=c_nig).value = (prefix + ("," if prefix else "") + f"{EXC_CODE_TRIP}={wt_night_added}")

            # Totals (LL/UL are already numeric 0 → naturally excluded)
            day_cols   = [get_column_letter(day_col(i))     for i in range(1, days_in_month + 1)]
            night_cols = [get_column_letter(day_col(i) + 1) for i in range(1, days_in_month + 1)]
            ws.cell(row=r, column=total_day_col).value   = f"=SUM({day_cols[0]}{r}:{day_cols[-1]}{r})"
            ws.cell(row=r, column=total_night_col).value = f"=SUM({night_cols[0]}{r}:{night_cols[-1]}{r})"
            ws.cell(row=r, column=total_col).value       = f"={get_column_letter(total_day_col)}{r}+{get_column_letter(total_night_col)}{r}"
            ws.cell(row=r, column=total_day_col).number_format   = "0.00"
            ws.cell(row=r, column=total_night_col).number_format = "0.00"
            ws.cell(row=r, column=total_col).number_format       = "0.00"

            # Rightmost Exceptions summary text (labels)
            if rowinfo.real:
                labels_all: List[str] = []
                for _d, flags in _exceptions_by_day(rowinfo.real, y, m).items():
                    labels_all.extend(flags.labels())
                if labels_all:
                    ws.cell(row=r, column=exc_col).value = ", ".join(sorted({lbl for lbl in labels_all}, key=str))

    filename = "Working_Hours.xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp
